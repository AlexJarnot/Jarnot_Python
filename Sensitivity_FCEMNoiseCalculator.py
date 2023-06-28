# Sensitivity and FC/EM Noise Calculator

import numpy as np
import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog
import xlsxwriter as xl
import pathlib
from openpyxl import load_workbook

root=tk.Tk()
root.resizable(width=False, height=False)
instradio=tk.IntVar()
filradio=tk.IntVar()
root.geometry('{}x{}'.format(1000, 300))
root.title('Sensitivity / FC & EM Noise Calculator')

test = tk.StringVar()
var = tk.IntVar()


def click_select():
    file_path = filedialog.askopenfilename()
    fileentry.delete(0,tk.END)
    fileentry.insert(0,file_path)
    loadupdate.config(text="")
    sensupdate.config(text="")
    fcemupdate.config(text="")
    
def click_load():
    global file_path
    file_path = fileentry.get()
    global worksheetname
    worksheetname = pathlib.Path(file_path)
    global workbook
    global worksheet
    global data
    data = pd.read_csv(file_path)
    workbook = xl.Workbook(str(worksheetname.parent) + r'\PostTestAnalysis.xlsx')
    loadupdate.config(text='File Loaded!')
    
def calc_sens():
    group1id = data.index[(data['Instr Scan Num'] == 2) & 
                          (data['Method Number'] == 10149) & 
                          (data['Electron Energy Setpoint'] == 40) & 
                          (data['Peak Top Widen State'] == 'On')].tolist()
    group2id = data.index[(data['Instr Scan Num'] == 2) & 
                          (data['Method Number'] == 10149) & 
                          (data['Electron Energy Setpoint'] == 40) & 
                          (data['Peak Top Widen State'] == 'Off')].tolist()
    group3id = data.index[(data['Instr Scan Num'] == 2) & 
                          (data['Method Number'] == 10149) & 
                          (data['Electron Energy Setpoint'] == 70) & 
                          (data['Peak Top Widen State'] == 'On')].tolist()
    group4id = data.index[(data['Instr Scan Num'] == 2) & 
                          (data['Method Number'] == 10149) & 
                          (data['Electron Energy Setpoint'] == 70) & 
                          (data['Peak Top Widen State'] == 'Off')].tolist()
    colindex = data.columns.get_loc('28')
    group1 = data.iloc[group1id[0]:(group1id[0] + 25), colindex]
    group2 = data.iloc[group2id[0]:(group2id[0] + 25), colindex]
    group3 = data.iloc[group3id[0]:(group3id[0] + 25), colindex]
    group4 = data.iloc[group4id[0]:(group4id[0] + 25), colindex]
    if var.get() == 1:
        basepres = 1
        testpres = 1
    else:
        basepres = startpresentry.get()
        testpres = endpresentry.get()
    deltapres = float(testpres) - float(basepres)
    ave1 = np.mean(group1)
    ave2 = np.mean(group2)
    ave3 = np.mean(group3)
    ave4 = np.mean(group4)
    sens1 = ave1/deltapres
    sens2 = ave2/deltapres
    sens3 = ave3/deltapres
    sens4 = ave4/deltapres
    
    worksheet = workbook.add_worksheet('Sensitivity')
    worksheet.write(0,0,'Base Pressure (torr)')
    worksheet.write(0,1,basepres)
    worksheet.write(1,0,'Test Pressure (torr)')
    worksheet.write(1,1,testpres)
    
    worksheet.write(5,0,'Sensitivity (40 EE, 200 EC, PTW ON)')
    worksheet.write(5,1,sens1)
    worksheet.write(3,2,'A/torr')
    worksheet.write(3,0,'Sensitivity (40 EE, 200 EC, PTW OFF)')
    worksheet.write(3,1,sens2)
    worksheet.write(4,2,'A/torr')
    worksheet.write(6,0,'Sensitivity (70 EE, 1000 EC, PTW ON)')
    worksheet.write(6,1,sens3)
    worksheet.write(5,2,'A/torr')
    worksheet.write(4,0,'Sensitivity (70 EE, 1000 EC, PTW OFF)')
    worksheet.write(4,1,sens4)
    worksheet.write(6,2,'A/torr')
    sensupdate.config(text='Sensitivity Calculated!')

def calc_fcem():
    group1id = data.index[(data['Instr Scan Num'] == 2) & 
                          (data['Method Number'] == 9726) & 
                          (data['Electron Multiplier State'] == 'Off')].tolist()
    group2id = data.index[(data['Instr Scan Num'] == 2) & 
                          (data['Method Number'] == 9826) & 
                          (data['Electron Multiplier State'] == 'Off')].tolist()
    group3id = data.index[(data['Instr Scan Num'] == 2) & 
                          (data['Method Number'] == 10226) & 
                          (data['Electron Multiplier State'] == 'Off')].tolist()
    group4id = data.index[(data['Instr Scan Num'] == 2) & 
                          (data['Method Number'] == 10426) & 
                          (data['Electron Multiplier State'] == 'Off')].tolist()
    group5id = data.index[(data['Instr Scan Num'] == 2) & 
                          (data['Method Number'] == 9726) & 
                          (data['Electron Multiplier State'] == 'On')].tolist()
    group6id = data.index[(data['Instr Scan Num'] == 2) & 
                          (data['Method Number'] == 9826) & 
                          (data['Electron Multiplier State'] == 'On')].tolist()
    group7id = data.index[(data['Instr Scan Num'] == 2) & 
                          (data['Method Number'] == 10226) & 
                          (data['Electron Multiplier State'] == 'On')].tolist()
    group8id = data.index[(data['Instr Scan Num'] == 2) & 
                          (data['Method Number'] == 10426) & 
                          (data['Electron Multiplier State'] == 'On')].tolist()
    colindex = data.columns.get_loc('5')
    group1 = data.iloc[group1id[0]:(group1id[0] + 25), colindex]
    group2 = data.iloc[group2id[0]:(group2id[0] + 25), colindex]
    group3 = data.iloc[group3id[0]:(group3id[0] + 25), colindex]
    group4 = data.iloc[group4id[0]:(group4id[0] + 25), colindex]
    group5 = data.iloc[group5id[0]:(group5id[0] + 25), colindex]
    group6 = data.iloc[group6id[0]:(group6id[0] + 25), colindex]
    group7 = data.iloc[group7id[0]:(group7id[0] + 25), colindex]
    group8 = data.iloc[group8id[0]:(group8id[0] + 25), colindex]
    ave1 = np.mean(group1)
    ave2 = np.mean(group2)
    ave3 = np.mean(group3)
    ave4 = np.mean(group4)
    ave5 = np.mean(group5)
    ave6 = np.mean(group6)
    ave7 = np.mean(group7)
    ave8 = np.mean(group8)
    sd1 = np.std(group1)
    sd2 = np.std(group2)
    sd3 = np.std(group3)
    sd4 = np.std(group4)
    sd5 = np.std(group5)
    sd6 = np.std(group6)
    sd7 = np.std(group7)
    sd8 = np.std(group8)
    
    worksheet = workbook.add_worksheet('FC EM Noise')
    worksheet.write(0,0,'FC Noise')
    worksheet.write(1,0,'Dwell (ms)')
    worksheet.write(1,1,'Moving Average (A)')
    worksheet.write(1,2,'Moving Standard Deviation')
    worksheet.write(2,0,'16')
    worksheet.write(3,0,'32')
    worksheet.write(4,0,'256')
    worksheet.write(5,0,'1024')
    
    worksheet.write(2,1,ave1)
    worksheet.write(2,2,sd1)
    worksheet.write(3,1,ave2)
    worksheet.write(3,2,sd2)
    worksheet.write(4,1,ave3)
    worksheet.write(4,2,sd3)
    worksheet.write(5,1,ave4)
    worksheet.write(5,2,sd4)
    
    worksheet.write(6,0,'EM Noise')
    worksheet.write(7,0,'Dwell (ms)')
    worksheet.write(7,1,'Moving Average (A)')
    worksheet.write(7,2,'Moving Standard Deviation')
    worksheet.write(8,0,'16')
    worksheet.write(9,0,'32')
    worksheet.write(10,0,'256')
    worksheet.write(11,0,'1024')
    
    worksheet.write(8,1,ave5)
    worksheet.write(8,2,sd5)
    worksheet.write(9,1,ave6)
    worksheet.write(9,2,sd6)
    worksheet.write(10,1,ave7)
    worksheet.write(10,2,sd7)
    worksheet.write(11,1,ave8)
    worksheet.write(11,2,sd8)
    fcemupdate.config(text='FC / EM Noise Calculated!')

def checkout():
    if var.get() == 1:
        startpresentry.configure(state='disabled')
        endpresentry.configure(state='disabled')
    else:
        startpresentry.configure(state='normal')
        endpresentry.configure(state='normal')
        
def close():
    workbook.close()
    newname = str(filename.get())
    oldfile = str(worksheetname.parent) + r'\PostTestAnalysis.xlsx'
    newfile = str(str(worksheetname.parent) + '/' + str(newname) + '_PostTestAnalysis.xlsx')
    os.replace(oldfile, newfile)
    root.destroy()
    book = load_workbook(newfile) 
    writer = pd.ExcelWriter(newfile, engine = 'openpyxl')
    writer.book = book
    pd.read_csv(file_path).to_excel(writer,'Raw Data')
    writer.save()


# Main Containers

headerframe=tk.Frame(width=1000, height=50, master=root, relief=tk.GROOVE, borderwidth=3)
inputframe=tk.Frame(width=1000, height=180, master=root, relief=tk.GROOVE, borderwidth=3)
outputframe=tk.Frame(root, width=1000, height=70)


# Lay out Main Frames

headerframe.grid(row=0,column=0, sticky='nsew')
inputframe.grid(row=1, column=0, sticky='nsew')
outputframe.grid(row=4, column=0, sticky='nsew')

inputframe.grid_propagate(0)
outputframe.grid_propagate(0)


# Header Banner

headerlabel=tk.Label(headerframe, text='Sensitivity / FC & EM Noise Calculator', font='Helvetica 12 bold')
headerlabel.grid(row=0, column=1)


# Create Input Widgets

selectlabel=tk.Label(inputframe, text='Select an Data File (.csv):')
fileentry=tk.Entry(inputframe, background='white', width=90)
dialogbutton=tk.Button(inputframe, text='Select File', padx=5, pady=5, command=click_select)
loadbutton=tk.Button(inputframe, text='Load', padx=5, pady=5, command=click_load)
calcsensbutton=tk.Button(inputframe, text='Calculate Sensitivity', padx=5, pady=5, command=calc_sens)
calcfcembutton=tk.Button(inputframe, text='Calculate FC / EM Noise', padx=5, pady=5, command=calc_fcem)
checklabel=tk.Label(inputframe, text='Use Pressures from Instrument Data?')
startpreslabel=tk.Label(inputframe, text='Base Pressure (torr):')
startpresentry=tk.Entry(inputframe, background='white', width=30)
endpreslabel=tk.Label(inputframe, text='Test Pressure (torr):')
endpresentry=tk.Entry(inputframe, background='white', width=30)
check=tk.Checkbutton(inputframe, onvalue=1, offvalue=0, command=checkout, variable=var)

filenameheader=tk.Label(outputframe, text='Set File Name Prefix:')
filename=tk.Entry(outputframe, background='white', width=30)
filenametxt=tk.Label(outputframe, text='_PostTestAnalysis.xlsx')
closebutton=tk.Button(outputframe, text='EXIT', padx=5, pady=5, command=close)

loadupdate=tk.Label(inputframe, text='')
sensupdate=tk.Label(inputframe, text='')
fcemupdate=tk.Label(inputframe, text='')


# Lay Out Input Widgets

selectlabel.grid(row=0, column=0, sticky='e')
fileentry.grid(row=0, column=1)
dialogbutton.grid(row=1, column=0, sticky='e')
loadbutton.grid(row=1, column=1, sticky='w')
calcsensbutton.grid(row=5, column=0, sticky='e')
calcfcembutton.grid(row=5, column=1, sticky='w')
checklabel.grid(row=2, column=0, sticky='e')
check.grid(row=2, column=1, sticky='w')
startpreslabel.grid(row=3, column=0, sticky='e')
startpresentry.grid(row=3, column=1, sticky='w')
endpreslabel.grid(row=4, column=0, sticky='e')
endpresentry.grid(row=4, column=1, sticky='w')

filenameheader.grid(row=0, column=0,  sticky='w')
filename.grid(row=1, column=0,  sticky='w')
filenametxt.grid(row=1, column=1,  sticky='w')
closebutton.grid(row=2, column=0, sticky='w')

loadupdate.grid(row=3, column=2, sticky='w')
sensupdate.grid(row=4, column=2, sticky='w')
fcemupdate.grid(row=5, column=2, sticky='w')


# Main Loop

root.mainloop()